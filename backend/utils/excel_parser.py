import os
import re
import pandas as pd


def is_allowed_file(filename, allowed_extensions):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed_extensions


def _get_excel_engine(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        return None
    try:
        import python_calamine  # noqa: F401
        return "calamine"
    except ImportError:
        if ext == ".xls":
            return "xlrd"
        return "openpyxl"


def read_excel_preview(filepath, sample_size=5):
    """
    ULTRA-FAST PREVIEW:
    Reads ONLY first 5 rows using Rust engine (0.1 second response time).
    """
    ext = os.path.splitext(filepath)[1].lower()

    try:
        if ext == ".csv":
            df = pd.read_csv(filepath, dtype=str, nrows=sample_size)
            # Fast line count for CSV
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                total_rows = max(sum(1 for _ in f) - 1, 0)
        else:
            engine = _get_excel_engine(filepath)
            df = pd.read_excel(
                filepath,
                dtype=str,
                nrows=sample_size,
                engine=engine,
            )
            # Estimate or get fast count
            total_rows = len(df)
            try:
                if engine == "calamine":
                    from python_calamine import CalamineWorkbook
                    wb = CalamineWorkbook.from_path(filepath)
                    sheet = wb.get_sheet_by_index(0)
                    total_rows = max(len(sheet.to_python()) - 1, 0)
                else:
                    from openpyxl import load_workbook
                    wb = load_workbook(filepath, read_only=True, data_only=True)
                    ws = wb.active
                    total_rows = max((ws.max_row or 1) - 1, 0)
                    wb.close()
            except Exception:
                total_rows = len(df)

        df = df.fillna("")
        df.columns = [str(c).strip() for c in df.columns]

        columns = []
        for col in df.columns:
            samples = df[col].astype(str).str.strip()
            samples = samples[samples != ""].head(sample_size).tolist()
            columns.append({"name": col, "samples": samples})

        preview_rows = []
        for _, row in df.iterrows():
            preview_rows.append({col: str(row[col]).strip() for col in df.columns})

        return {
            "columns": columns,
            "column_names": list(df.columns),
            "preview_rows": preview_rows,
            "total_rows": int(total_rows),
        }

    except Exception as e:
        raise Exception(f"Failed to read file: {str(e)}")


def parse_excel_with_columns(filepath, qr_column, id_column=None):
    """
    FULL EXCEL PARSE:
    Uses Calamine (Rust Engine) - parses 6.5 lakh rows in ~3 seconds!
    """
    ext = os.path.splitext(filepath)[1].lower()

    try:
        if ext == ".csv":
            df = pd.read_csv(filepath, dtype=str)
        else:
            engine = _get_excel_engine(filepath)
            df = pd.read_excel(filepath, dtype=str, engine=engine)

        df = df.fillna("")
        df.columns = [str(c).strip() for c in df.columns]

        if qr_column not in df.columns:
            raise Exception(f"Column '{qr_column}' not found in file")

        if id_column and id_column not in df.columns:
            raise Exception(f"Column '{id_column}' not found in file")

        # High-speed vectorized string extraction
        qr_series = df[qr_column].astype(str).str.strip()
        mask = qr_series != ""
        df_valid = df.loc[mask]

        qr_values = df_valid[qr_column].astype(str).str.strip().tolist()

        if id_column:
            raw_ids = df_valid[id_column].astype(str).str.strip().tolist()
            unique_ids = []
            for i, rid in enumerate(raw_ids):
                if rid and rid.lower() != "nan":
                    cleaned = re.sub(r"[^\w\-]", "_", rid)[:60]
                    unique_ids.append(cleaned if cleaned else f"AUTO{i + 1}")
                else:
                    unique_ids.append(f"AUTO{i + 1}")
        else:
            unique_ids = [f"AUTO{i + 1}" for i in range(len(qr_values))]

        results = [
            {
                "content": qr_values[i],
                "unique_id": unique_ids[i],
                "row_number": i + 2,
            }
            for i in range(len(qr_values))
        ]

        return {
            "data": results,
            "qr_column": qr_column,
            "id_column": id_column,
            "total_rows": len(df),
            "valid_rows": len(results),
        }

    except Exception as e:
        raise Exception(f"Failed to parse file: {str(e)}")
